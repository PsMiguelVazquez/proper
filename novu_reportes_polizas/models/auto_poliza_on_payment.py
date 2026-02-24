# -*- coding: utf-8 -*-
import io
import os
import base64
import logging
import datetime
import pytz
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from odoo import models, api, tools, fields

_logger = logging.getLogger(__name__)


class AccountMove(models.Model):
    _inherit = 'account.move'
    poliza_generada = fields.Boolean(string='Póliza generada', default=False, copy=False)
    poliza_provision_generada = fields.Boolean(string='Póliza Provisión Generada', default=False, copy=False)
    poliza_pago_generada = fields.Boolean(string='Póliza de Pago Generada', default=False, copy=False)   
    def _generate_poliza_attachment(self):

        _logger.error(f"Llegue a generar")
        if self.env.context.get('skip_poliza'):
            return
        
        self.ensure_one()

        addons_paths = tools.config['addons_path'].split(',')
        module_name = 'novu_reportes_polizas'
        file_relative_path = 'static/layouts/Layout_poliza_diario.xlsx'
        path = False
        for addons_path in addons_paths:
            test_path = os.path.join(addons_path.strip(), module_name, file_relative_path)
            if os.path.exists(test_path):
                path = test_path
                break

        if not path:
            _logger.error('No se encontró la plantilla %s en los addons_path', file_relative_path)
            return False

        workbook = load_workbook(filename=path)
        sheet = workbook.active

        # Valores para encabezado
        tz = pytz.timezone(self.env.user.tz or 'America/Mexico_City')
        timestamp2 = datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')
        timestamp = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')

        date_use = self.invoice_date or self.date or fields.Date.context_today(self)
        if isinstance(date_use, str):
            date_obj = datetime.datetime.strptime(date_use, '%Y-%m-%d').date()
        else:
            date_obj = date_use

        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        mes_nombre = meses.get(date_obj.month, "")

        folio_poliza = self.env['ir.sequence'].next_by_code('poliza.diario.seq') or '/'

        # Escribir encabezados en plantilla (mismas celdas que el wizard)
        def safe_write(sheet, cell_ref, value):
            cell = sheet[cell_ref]
            if isinstance(cell, MergedCell):
                for merge_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        master_cell = sheet.cell(merge_range.min_row, merge_range.min_col)
                        master_cell.value = value
                        return
                raise ValueError("No se encontró celda combinada para %s" % cell_ref)
            else:
                cell.value = value

        safe_write(sheet, 'B2', f"{folio_poliza}")
        safe_write(sheet, 'B3', f"{timestamp2}")
        safe_write(sheet, 'B4', f"{mes_nombre.upper()} EJERCICIO: {date_obj.year}")
        safe_write(sheet, 'B5', self.name or self.invoice_origin or '')

        dia = date_obj.day
        row = 8

        #Obtener las lineas de la factura
        invoice_lines = self.line_ids

        #obener lineas consiliables
        reconcilable = invoice_lines.filtered(
            lambda l: l.account_id.internal_type in ('receivable', 'payable')
        )
        
        #Obtener lo conciliado
        related_lines = self.env['account.move.line']
        related_lines |= invoice_lines
    
        for line in reconcilable:
            related_lines |= line.matched_debit_ids.mapped('debit_move_id')
            related_lines |= line.matched_credit_ids.mapped('credit_move_id')
    
        #Lineas extras como pagos
        all_moves = related_lines.mapped('move_id')
        all_lines = all_moves.mapped('line_ids')
    
        #Ordenar las lineas
        all_lines = all_lines.sorted(
            lambda l: (l.date, l.move_id.id, l.id)
        )
    
        #Escribir las lineas en el reporte
        row = 9
        suma_debitos = 0.0
        suma_creditos = 0.0
        last_move = False
    
        for line in all_lines:
            #Línea en blanco entre asientos
            if last_move and last_move != line.move_id:
                row += 1
    
            sheet[f"A{row}"] = line.account_id.code or ''
            sheet[f"B{row}"] = line.account_id.name or ''
            sheet[f"C{row}"] = line.move_id.name or ''
            sheet[f"D{row}"] = dia
    
            if line.debit:
                sheet[f"G{row}"] = float(line.debit)
                sheet[f"G{row}"].number_format = '#,##0.00'
                suma_debitos += line.debit
    
            if line.credit:
                sheet[f"H{row}"] = float(line.credit)
                sheet[f"H{row}"].number_format = '#,##0.00'
                suma_creditos += line.credit                
            last_move = line.move_id
            row += 1
    
        sheet[f"F{row}"] = "Totales"
        sheet[f"G{row}"] = suma_debitos
        sheet[f"H{row}"] = suma_creditos
        sheet[f"G{row}"].number_format = '#,##0.00'
        sheet[f"H{row}"].number_format = '#,##0.00'

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        attachment = self.env['ir.attachment'].create({
            'name': f'Poliza_de_diario_{self.name}_{timestamp}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'account.move',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        output.close()

        self.poliza_generada = True

        _logger.info(f"Póliza generada correctamente (attachment id={attachment.id}) para {self.name}")
        return True


    def descarga_poliza(self):
        self.ensure_one()
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'account.move'),
            ('res_id', '=', self.id),
            ('name', 'ilike', 'Poliza_de_diario_')
        ], order='create_date desc', limit=1)
    
        if not attachment:
            return False
    
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def generate_poliza_provision_attachment(self):

        self.ensure_one()
        addons_paths = tools.config['addons_path'].split(',')
        module_name = 'novu_reportes_polizas'
        file_relative_path = 'static/layouts/Layout_poliza_diario.xlsx'
        path = False
        for addons_path in addons_paths:
            test_path = os.path.join(addons_path.strip(), module_name, file_relative_path)
            if os.path.exists(test_path):
                path = test_path
                break

        if not path:
            _logger.error('No se encontró la plantilla %s', file_relative_path)
            return False

        workbook = load_workbook(filename=path)
        sheet = workbook.active


        tz = pytz.timezone(self.env.user.tz or 'America/Mexico_City')
        timestamp_str = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')
        
        date_use = self.invoice_date or self.date or fields.Date.context_today(self)
        if isinstance(date_use, str):
            date_obj = datetime.datetime.strptime(date_use, '%Y-%m-%d').date()
        else:
            date_obj = date_use

        meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
                 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
        mes_nombre = meses.get(date_obj.month, "")
        

        folio_poliza = self.name or '/' 


        def safe_write(sheet, cell_ref, value):
            cell = sheet[cell_ref]
            if isinstance(cell, MergedCell):
                for merge_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        sheet.cell(merge_range.min_row, merge_range.min_col).value = value
                        return
            else:
                cell.value = value

        safe_write(sheet, 'B2', f"{folio_poliza} (DIARIO)") # Distintivo visual
        safe_write(sheet, 'B3', datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S'))
        safe_write(sheet, 'B4', f"{mes_nombre.upper()} EJERCICIO: {date_obj.year}")
        safe_write(sheet, 'B5', self.name or '')


        all_lines = self.line_ids.sorted(key=lambda l: (l.account_id.code, l.id))

        row = 9
        suma_debitos = 0.0
        suma_creditos = 0.0
        dia = date_obj.day

        for line in all_lines:
            sheet[f"A{row}"] = line.account_id.code or ''
            sheet[f"B{row}"] = line.account_id.name or ''
            sheet[f"C{row}"] = line.move_id.name or ''
            sheet[f"D{row}"] = dia

            debit = float(line.debit)
            credit = float(line.credit)

            if debit:
                sheet[f"G{row}"] = debit
                sheet[f"G{row}"].number_format = '#,##0.00'
                suma_debitos += debit

            if credit:
                sheet[f"H{row}"] = credit
                sheet[f"H{row}"].number_format = '#,##0.00'
                suma_creditos += credit

            row += 1


        sheet[f"F{row}"] = "Totales"
        sheet[f"G{row}"] = suma_debitos
        sheet[f"H{row}"] = suma_creditos
        sheet[f"G{row}"].number_format = '#,##0.00'
        sheet[f"H{row}"].number_format = '#,##0.00'

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        

        filename = f'Poliza_Diario_{self.name.replace("/", "_")}_{timestamp_str}.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'account.move',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        output.close()
        
        self.poliza_provision_generada = True
        return True

    def descargar_poliza_provision(self):

        self.ensure_one()
        attachment = self.env['ir.attachment'].search([ ('res_model', '=', 'account.move'), ('res_id', '=', self.id), ('name', 'ilike', 'Poliza_Diario_')], order='create_date desc', limit=1)
        
        if not attachment:
            self.generate_poliza_provision_attachment()
            attachment = self.env['ir.attachment'].search([ ('res_model', '=', 'account.move'), ('res_id', '=', self.id),('name', 'ilike', 'Poliza_Diario_') ], order='create_date desc', limit=1)

        if attachment:
            return {'type': 'ir.actions.act_url','url': f'/web/content/{attachment.id}?download=true','target': 'self',}
        return False

    
    def descargar_poliza_pago(self):

        self.ensure_one()
        if self.move_type == "out_invoice":
            attachment = self.env['ir.attachment'].search([
                ('res_model', '=', 'account.move'),
                ('res_id', '=', self.id),
                ('name', 'ilike', 'Poliza_Ingreso_')
            ], order='create_date desc', limit=1)
        else:
            attachment = self.env['ir.attachment'].search([
                ('res_model', '=', 'account.move'),
                ('res_id', '=', self.id),
                ('name', 'ilike', 'Poliza_Egreso_')
            ], order='create_date desc', limit=1)
        
        if not attachment:

            if self.payment_state in ('paid', 'in_payment'):
                _logger.error("está pagadaasdasad")
                self._generate_poliza_pago_attachment()
                if self.move_type == "out_invoice":
                    attachment = self.env['ir.attachment'].search([ ('res_model', '=', 'account.move'), ('res_id', '=', self.id),('name', 'ilike', 'Poliza_Ingreso_') ], order='create_date desc', limit=1)
                else:
                    attachment = self.env['ir.attachment'].search([ ('res_model', '=', 'account.move'), ('res_id', '=', self.id),('name', 'ilike', 'Poliza_Egreso_') ], order='create_date desc', limit=1)
        if attachment:
            return {
                        'type': 'ir.actions.act_url',
                        'url': f'/web/content/{attachment.id}?download=true',
                        'target': 'self',
                    }
            
            # if not attachment:
            #     return {
            #         'type': 'ir.actions.client',
            #         'tag': 'display_notification',
            #         'params': {
            #             'title': 'Sin Póliza',
            #             'message': 'No se encontró una póliza de pago (la factura quizás no está pagada aún).',
            #             'type': 'warning',
            #             'sticky': False,
            #         }
            #     }

        


    def _generate_poliza_pago_attachment(self):
        self.ensure_one()
        _logger.info(f"--- GENERANDO POLIZA PAGO PERFECTA (IVA COMPLETO) PARA: {self.name} ---")

        addons_paths = tools.config['addons_path'].split(',')
        module_name = 'novu_reportes_polizas'

        if self.move_type == "out_invoice":
            file_relative_path = 'static/layouts/Layout_poliza_ingreso.xlsx'
        else:
            file_relative_path = 'static/layouts/Layout_poliza_egreso.xlsx'
            
        path = False
        for addons_path in addons_paths:
            test_path = os.path.join(addons_path.strip(), module_name, file_relative_path)
            if os.path.exists(test_path):
                path = test_path
                break

        if not path:
            _logger.error('No se encontró la plantilla %s', file_relative_path)
            return False

        workbook = load_workbook(filename=path)
        sheet = workbook.active

        tz = pytz.timezone(self.env.user.tz or 'America/Mexico_City')
        timestamp_str = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')

        invoice_lines = self.line_ids.filtered(
            lambda l: l.account_id.internal_type in ('receivable', 'payable')
        )
        
        partials = self.env['account.partial.reconcile']
        for line in invoice_lines:
            partials |= line.matched_debit_ids
            partials |= line.matched_credit_ids
        
        related_moves = (partials.mapped('debit_move_id.move_id') | partials.mapped('credit_move_id.move_id'))
        

        cbmx_moves = self.env['account.move']

        if partials:
            cbmx_moves |= self.env['account.move'].search([
                ('tax_cash_basis_rec_id', 'in', partials.ids),
                ('state', '=', 'posted')
            ])

        cbmx_moves |= self.env['account.move'].search([
            ('tax_cash_basis_origin_move_id', '=', self.id),
            ('state', '=', 'posted')
        ])

        if self.tax_cash_basis_created_move_ids:
            cbmx_moves |= self.tax_cash_basis_created_move_ids.filtered(lambda m: m.state == 'posted')


        payment_moves = related_moves - self
        payment_moves = payment_moves - cbmx_moves

        if not payment_moves:
             _logger.warning("No se detectaron pagos bancarios principales.")

     
        if payment_moves:
            last_payment_date = max(payment_moves.mapped('date'))
        else:
            last_payment_date = fields.Date.context_today(self)
            
        meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
                 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
        mes_nombre = meses.get(last_payment_date.month, "")

        def safe_write(sheet, cell_ref, value):
            cell = sheet[cell_ref]
            if isinstance(cell, MergedCell):
                for merge_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        sheet.cell(merge_range.min_row, merge_range.min_col).value = value
                        return
            else:
                cell.value = value

        nombres_pagos = ", ".join(payment_moves.mapped('name'))
        tipo_poliza = "INGRESO" if self.move_type == "out_invoice" else "EGRESO"

        safe_write(sheet, 'B2', f"{nombres_pagos} ({tipo_poliza})") 
        safe_write(sheet, 'B3', datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S'))
        safe_write(sheet, 'B4', f"{mes_nombre.upper()} EJERCICIO: {last_payment_date.year}")
        safe_write(sheet, 'B5', f"Pago de: {self.name}")

        row = 9
        
 
        suma_debitos = 0.0
        suma_creditos = 0.0
        
        payment_lines = payment_moves.mapped('line_ids').sorted(key=lambda l: (l.move_id.date, l.move_id.name, l.id))

        for line in payment_lines:
            sheet[f"A{row}"] = line.account_id.code or ''
            sheet[f"B{row}"] = line.account_id.name or ''
            sheet[f"C{row}"] = line.move_id.name or '' 
            sheet[f"D{row}"] = line.date.day

            debit = float(line.debit)
            credit = float(line.credit)

            if debit:
                sheet[f"G{row}"] = debit
                sheet[f"G{row}"].number_format = '#,##0.00'
                suma_debitos += debit

            if credit:
                sheet[f"H{row}"] = credit
                sheet[f"H{row}"].number_format = '#,##0.00'
                suma_creditos += credit

            row += 1

        sheet[f"F{row}"] = "Totales Pago"
        sheet[f"G{row}"] = suma_debitos
        sheet[f"H{row}"] = suma_creditos
        sheet[f"G{row}"].number_format = '#,##0.00'
        sheet[f"H{row}"].number_format = '#,##0.00'
        
        row += 3 


        if cbmx_moves:
            safe_write(sheet, f"B{row}", "--- MOVIMIENTOS DE IMPUESTOS (BASE EFECTIVO) ---")
            row += 1

            suma_tax_debitos = 0.0
            suma_tax_creditos = 0.0
            
            # === CAMBIO DEFINITIVO ===
            # Filtramos EXCLUYENDO las cuentas que empiezan con '8'.
            # En la contabilidad MX, las cuentas 8xx son "Cuentas de Orden" usadas para la Base Imponible (DIOT).
            # Las cuentas de IVA real son 1xx o 2xx.
            tax_lines = cbmx_moves.mapped('line_ids').filtered(
                lambda l: not l.account_id.code.startswith('8')
            ).sorted(key=lambda l: (l.account_id.code, l.id))

            for line in tax_lines:
                sheet[f"A{row}"] = line.account_id.code or ''
                sheet[f"B{row}"] = line.account_id.name or ''
                
                ref = line.move_id.name or ''
                sheet[f"C{row}"] = f"{ref} (IMP)" 
                sheet[f"D{row}"] = line.date.day

                debit = float(line.debit)
                credit = float(line.credit)

                if debit:
                    sheet[f"G{row}"] = debit
                    sheet[f"G{row}"].number_format = '#,##0.00'
                    suma_tax_debitos += debit

                if credit:
                    sheet[f"H{row}"] = credit
                    sheet[f"H{row}"].number_format = '#,##0.00'
                    suma_tax_creditos += credit

                row += 1
            
            sheet[f"F{row}"] = "Totales Impuestos"
            sheet[f"G{row}"] = suma_tax_debitos
            sheet[f"H{row}"] = suma_tax_creditos
            sheet[f"G{row}"].number_format = '#,##0.00'
            sheet[f"H{row}"].number_format = '#,##0.00'

        # ... [RESTO DEL CÓDIGO DE GUARDADO IGUAL] ...
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'Poliza_{tipo_poliza.title()}_{self.name.replace("/", "_")}_{timestamp_str}.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'account.move',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        output.close()

        self.poliza_pago_generada = True
        return True

class StockPicking(models.Model):
    _inherit = 'stock.picking'

    def action_descargar_poliza_diario(self):
        """
        Genera y descarga la póliza contable asociada a este traslado
        directamente desde el modelo stock.picking.
        """
        self.ensure_one()

   
        stock_moves = False
        if hasattr(self, 'move_lines') and self.move_lines:
            stock_moves = self.move_lines
        elif hasattr(self, 'move_ids_without_package') and self.move_ids_without_package:
            stock_moves = self.move_ids_without_package
            
        if not stock_moves:
             raise UserError(_('No se encontraron líneas de productos en este traslado.'))

        account_moves = stock_moves.mapped('account_move_ids').filtered(lambda m: m.state == 'posted')

        if not account_moves:
            raise UserError(_(
                'No se encontró una Póliza Contable publicada para este traslado.\n'
                'Asegúrate de que el traslado esté en estado "Realizado" (Done) y que '
                'la categoría del producto tenga valoración automática.'
            ))
        

        poliza_move = account_moves[0]


        attachment_name_pattern = f'Poliza_Stock_{self.name.replace("/", "_")}'
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'account.move'),
            ('res_id', '=', poliza_move.id),
            ('name', 'ilike', attachment_name_pattern)
        ], order='create_date desc', limit=1)

        if attachment:
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }

        addons_paths = tools.config['addons_path'].split(',')
        module_name = 'novu_reportes_polizas'
        file_relative_path = 'static/layouts/Layout_poliza_diario.xlsx'
        path = False
        for addons_path in addons_paths:
            test_path = os.path.join(addons_path.strip(), module_name, file_relative_path)
            if os.path.exists(test_path):
                path = test_path
                break

        if not path:
            raise UserError(_(f'No se encontró la plantilla {file_relative_path} en el servidor.'))

        workbook = load_workbook(filename=path)
        sheet = workbook.active

        tz = pytz.timezone(self.env.user.tz or 'America/Mexico_City')
        timestamp_str = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')
        
        date_obj = poliza_move.date or fields.Date.context_today(self)
        
        meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
                 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
        mes_nombre = meses.get(date_obj.month, "")

        def safe_write(sheet, cell_ref, value):
            cell = sheet[cell_ref]
            if isinstance(cell, MergedCell):
                for merge_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        sheet.cell(merge_range.min_row, merge_range.min_col).value = value
                        return
            else:
                cell.value = value


        safe_write(sheet, 'B2', f"{self.name} (STOCK)")
        safe_write(sheet, 'B3', datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S'))
        safe_write(sheet, 'B4', f"{mes_nombre.upper()} EJERCICIO: {date_obj.year}")
        safe_write(sheet, 'B5', f"Ref Asiento: {poliza_move.name}")

        all_lines = poliza_move.line_ids.sorted(key=lambda l: (l.account_id.code, l.id))

        row = 9
        suma_debitos = 0.0
        suma_creditos = 0.0
        dia = date_obj.day

        for line in all_lines:
            sheet[f"A{row}"] = line.account_id.code or ''
            sheet[f"B{row}"] = line.account_id.name or ''
            sheet[f"C{row}"] = line.move_id.name or ''
            sheet[f"D{row}"] = dia

            debit = float(line.debit)
            credit = float(line.credit)

            if debit:
                sheet[f"G{row}"] = debit
                sheet[f"G{row}"].number_format = '#,##0.00'
                suma_debitos += debit

            if credit:
                sheet[f"H{row}"] = credit
                sheet[f"H{row}"].number_format = '#,##0.00'
                suma_creditos += credit

            row += 1

        sheet[f"F{row}"] = "Totales Stock"
        sheet[f"G{row}"] = suma_debitos
        sheet[f"H{row}"] = suma_creditos
        sheet[f"G{row}"].number_format = '#,##0.00'
        sheet[f"H{row}"].number_format = '#,##0.00'

   
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'{attachment_name_pattern}_{timestamp_str}.xlsx'


        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'account.move',
            'res_id': poliza_move.id, 
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        output.close()

        _logger.info(f"Póliza de Stock generada desde Picking {self.name} para el asiento {poliza_move.name}")

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
        
class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'
    
    # def reconcile(self):
    #     _logger.error("ENTRO A reconcile")

    #     moves = self.mapped('move_id')
    #     res = super().reconcile()

    #     for move in moves:
    #         if (
    #             move.move_type in ('out_invoice', 'in_invoice')
    #             and move.payment_state == 'paid'
    #             and not move.poliza_generada
    #         ):
    #             try:
    #                 move._generate_poliza_attachment()
    #                 move.poliza_generada = True
    #             except Exception:
    #                 _logger.exception(f'Error generando póliza para factura {move.name}')

    #     return res

    def remove_move_reconcile(self):
        moves = self.mapped('move_id')
        res = super().remove_move_reconcile()

        for move in moves:
            if move.payment_state != 'paid':
                move.poliza_generada = False
                #Eliminar la poliza previa para poner la nueva cuando se concilie nuevamente
                attachments = self.env['ir.attachment'].search([
                ('res_model', '=', 'account.move'),
                ('res_id', '=', move.id),
                ('name', 'ilike', 'Poliza_de_diario_'),
            ])

            if attachments:
                _logger.info(
                    f'Eliminando {len(attachments)} póliza(s) de la factura {move.name} por desconciliación')
                attachments.sudo().unlink()

        return res
        