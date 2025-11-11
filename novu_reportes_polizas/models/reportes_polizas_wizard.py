# -*- coding: utf-8 -*-

import json
import io
import base64
import xlsxwriter
import datetime
from pytz import timezone
from odoo.tools import date_utils
from odoo.tools.misc import xlsxwriter
from odoo import fields, models, tools, api, _
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import numbers

import os
import logging
import pytz
_logger = logging.getLogger(__name__)

class ReportesPolizasWizard(models.TransientModel):
    _name = "reportes.polizas.wizard"
    _description = 'Wizard de filtros para reportes de polizas'

    name = fields.Char(string='Name', default="")
    descripcion = fields.Text(string='Descripción de poliza')
    #cliente = fields.Many2one('res.partner', string='Cliente')
    date_start = fields.Date(string='Date Start', required=True, default=datetime.datetime.now().strftime('%Y-%m-01'))


    def action_report(self):
        user_tz = self.env.user.tz or 'UTC'
        _logger.error(f"user_tz {user_tz}")
        tz = timezone(user_tz)

        if not self.date_start:
            raise UserError("Debe Proporcionar una fecha")
            
        descripción = self.descripcion
        fecha_inicio_local = tz.localize(datetime.datetime.combine(self.date_start, datetime.time.min))
        
        #Convertir a UTC (lo que realmente está en la base de datos)
        fecha_inicio = fecha_inicio_local.astimezone(datetime.timezone.utc)
        
        _logger.error(f"fecha_inicio {fecha_inicio} ")

        

        addons_paths = tools.config['addons_path'].split(',')
        module_name = 'novu_reportes_polizas'
        file_relative_path = 'static/layouts/Layout_poliza_diario.xlsx'
        path = False
               
        
        for addons_path in addons_paths:
            test_path = os.path.join(addons_path.strip(), module_name, file_relative_path)
            if os.path.exists(test_path):
                path = test_path
                break

        if not path:
            raise UserError(_("No se encontró la plantilla layout_kardex.xlsx"))

        #Abrir el archivo plantilla
        workbook = load_workbook(filename=path)
        sheet = workbook.active  # o selecciona por nombre: workbook['NombreHoja']

        tz = pytz.timezone('America/Mexico_City')
        timestamp = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')
        timestamp2 = datetime.datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')
        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        
        mes_nombre = meses[self.date_start.month]

        folio_poliza = self.env['ir.sequence'].next_by_code('poliza.diario.seq') or '/'
        
        self.safe_write(sheet, 'B2', f"{folio_poliza}")
        self.safe_write(sheet, 'B3', f"{timestamp2}")
        self.safe_write(sheet, 'B4', f"{mes_nombre.upper()} EJERCICIO: {self.date_start.year}")
        
        self.safe_write(sheet, 'B5', f"{descripción}")

        dia = self.date_start.day
        row = 8

        asientos = self.env['account.move.line'].search([
            ("parent_state", "=", "posted"),
            ("move_id.move_type","=","out_invoice"),
            ("account_id.code","not ilike","115.01"),
            ("account_id.code","not ilike","115-01"), 
            ("account_id.code","not ilike","107.05."),
            ("date", '=', fecha_inicio),
        ], order='date, move_id desc')

        asientos_clientes = asientos.filtered(lambda m: m.account_id.code.startswith('105'))
        asientos_otros = asientos.filtered(lambda m: not m.account_id.code.startswith('105'))
        mov_ant = ""
        suma_debitos = 0
        suma_creditos_iva = 0
        suma_creditos_ventas = 0
        total_debitos = 0
        total_creditos = 0
        suma_creditos = 0
        for mov in asientos_clientes:
            factura = mov.move_id.name
            if mov_ant == factura:
                row += 1
            else:
                mov_ant = factura
                row += 2

            if mov.debit > 0:
                sheet[f"A{row}"] = mov.account_id.code
                sheet[f"B{row}"] = mov.partner_id.name
                sheet[f"C{row}"] = mov.move_id.name
                sheet[f"D{row}"] = dia
                #sheet[f"G{row}"] = float(mov.debit or 0.0)
                cell_debit = sheet[f"G{row}"]
                cell_debit.value = float(mov.debit or 0.0)
                cell_debit.number_format = '#,##0.00'
                suma_debitos += mov.debit
            # elif mov.credit > 0:
            #     suma_creditos += mov.credit
                
        row += 1
        resumen_por_cuenta = {}
        for mov in asientos_otros:
            code = mov.account_id.code
            name = mov.account_id.name or ""
            resumen_por_cuenta.setdefault(code, {"name": name, "debit": 0.0, "credit": 0.0})
            resumen_por_cuenta[code]["debit"] += float(mov.debit or 0.0)
            resumen_por_cuenta[code]["credit"] += float(mov.credit or 0.0)
            suma_creditos += mov.credit
        
        for code, datos in sorted(resumen_por_cuenta.items()):
            total_debito = datos["debit"]
            total_credito = datos["credit"]
            dia = self.date_start.day
            sheet[f"A{row}"] = code
            sheet[f"B{row}"] = datos["name"]
            sheet[f"D{row}"] = dia
            
            if total_debito > total_credito:
                cell_total_debit = sheet[f"G{row}"]
                cell_total_debit.value = float(total_debito - total_credito or 0.0)
                cell_total_debit.number_format = '#,##0.00'
                #sheet[f"G{row}"] = total_debito - total_credito  # Cargo
            elif total_credito > total_debito:
                cell_total_credit = sheet[f"H{row}"]
                cell_total_credit.value = float(total_credito - total_debito or 0.0)
                cell_total_credit.number_format = '#,##0.00'
                #sheet[f"H{row}"] = total_credito - total_debito  # Abono
            else:
                sheet[f"G{row}"] = 0
                sheet[f"H{row}"] = 0
            row +=2

        total_debitos += total_debito
        total_creditos += total_credito

        sheet[f"F{row}"] = "Totales"
        cell_suma_total_debit = sheet[f"G{row}"]
        cell_suma_total_debit.value = float(suma_debitos or 0.0)
        cell_suma_total_debit.number_format = '#,##0.00'
        #sheet[f"G{row}"] = f"{suma_debitos}"

        cell_suma_total_credit = sheet[f"H{row}"]
        cell_suma_total_credit.value = float(suma_creditos or 0.0)
        cell_suma_total_credit.number_format = '#,##0.00'
        #sheet[f"H{row}"] = f"{suma_creditos}"
    

        tz = pytz.timezone('America/Mexico_City')
        timestamp = datetime.datetime.now(tz).strftime('%Y-%m-%d_%H_%M_%S')
            
        #Guardar el archivo en memoria
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        #Crear adjunto
        attachment = self.env['ir.attachment'].create({
            'name': f'Poliza_de_diario_{timestamp}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'reportes.polizas.wizard',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        output.close()

        #Descargar
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }



    def safe_write(self, sheet, cell_ref, value):
        """Escribe en una celda, incluso si está combinada."""
        cell = sheet[cell_ref]
        if isinstance(cell, MergedCell):
            # Buscar la celda principal de ese merge
            for merge_range in sheet.merged_cells.ranges:
                if cell.coordinate in merge_range:
                    # El primer (min_row, min_col) es la celda principal
                    master_cell = sheet.cell(merge_range.min_row, merge_range.min_col)
                    master_cell.value = value
                    return
            raise ValueError(f"No se encontró celda combinada para {cell_ref}")
        else:
            # Si no es merged, escribir normalmente
            cell.value = value
    